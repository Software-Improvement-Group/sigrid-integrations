from enum import Enum
from functools import singledispatch
from typing import Dict, Any, List

from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook


@singledispatch
def parseType(inp):
    return inp


@parseType.register
def parseTypeStr(inp: list):
    if len(inp) == 0:
        return "[]"
    else:
        return '[\'' + "\', \'".join(inp) + "\']"


def set_up_worksheet(xlsx_file):
    wb = load_workbook(xlsx_file)
    wb.active = 0
    header_row = [x.value for x in wb.active[1]]
    rows = wb.active.rows
    next(rows)
    return header_row, rows


class ExcelTypes(Enum):
    BOOL = "bool"
    STRING_ARRAY = "string_array"
    INT = "int"
    STRING = "string"


def _checkbool(input) -> bool:
    if isinstance(input, bool):
        return input
    if isinstance(input, str):
        return input.upper() == "TRUE"
    raise TypeError(f'Unsupported type provided, input was {input}')


def _checkint(input: str) -> int:
    out = None
    if not isinstance(input, int):
        try:
            out = int(input)
        except (ValueError, TypeError):
            # in this case there was probably an empty line
            print(f'Something went wrong parsing an int, probably empty line.  Value was {input}')
    else:
        out = input
    return out


def _checkArr(input: str) -> List[str]:
    if isinstance(input, str):
        if input == '[]' or input == '':
            return []
        else:
            return [x.strip() for x in input.replace("[", '').replace("]", "")
            .replace('\'', '').replace("\"", '').split(',')]
    else:
        raise ValueError(f'Unsupported type, got {input} of type {type(input)}')


# Reads from an already-zipped row and applies a type mapping.  Then returns as a Dict
def read_row_as_type(row_dict: Dict[str, str], type_mapping: Dict[str, ExcelTypes]) -> Dict[str, Any]:
    out_dict = {}

    type_validators = {ExcelTypes.BOOL: _checkbool, ExcelTypes.STRING_ARRAY: _checkArr, ExcelTypes.INT: _checkint,
                     ExcelTypes.STRING: lambda x: x}

    for key, excel_type in type_mapping.items():
        if key not in row_dict:
            print(f"warning: missing key {key}")
            continue
        try:
            out_dict[key] = type_validators[excel_type](row_dict[key])
        except KeyError:
            print(f'Did not find type {excel_type} in type lookup table.')
            out_dict[key] = row_dict[key]
    return out_dict


def read_rows_as_type(header_row, rows, type_mapping: Dict[str, ExcelTypes]) -> List[Dict[str, Any]]:
    """
    Receives a header_row, a list of rows, and a dict mapping keys in the header_row to type strings.
    :type type_mapping: Dict[str, str]
    """
    out_rows = []
    for row in rows:
        row_data = [x.value for x in row]
        data = dict(zip(header_row, row_data))
        out_rows.append(read_row_as_type(data, type_mapping))

    return out_rows


def load_from_excel_as_type(xlsx_file, type_mapping: Dict[str, ExcelTypes]) -> List[Dict[str, Any]]:
    header_row, rows = set_up_worksheet(xlsx_file)
    return read_rows_as_type(header_row, rows, type_mapping)

def create_template_from_array(header_row: List[str], output_file: str):
    wb = Workbook()
    mainsheet = wb.active
    mainsheet.title = f'system_data'
    mainsheet.append(header_row)
    wb.save(output_file)
    wb.close()


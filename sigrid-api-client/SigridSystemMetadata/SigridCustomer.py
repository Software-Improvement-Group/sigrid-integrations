import csv
import json
import time
from dataclasses import dataclass
from typing import Dict, Optional, List

from openpyxl import Workbook

from SigridRest.SigridGetMaintainabilityCommand import SigridGetMaintainabilityCommand
from SigridRest.SigridGetMetadataCommand import SigridGetMetadataCommand
from SigridRest.SigridGetArchitectureJSONCommand import SigridGetArchitectureJSONCommand
from SigridRest.SigridGetMaintainabilityJSONCommand import SigridGetMaintainabilityJSONCommand
from SigridRest.SigridOnboardQSMCommand import SigridOnboardQSMCommand
from SigridSystemMetadata.SigridMetadata import SigridMetadata
from SigridSystemMetadata.SigridObjectives import SigridObjectives
from SigridSystemMetadata.SigridSecurityResults import SigridSecurityResults
from SigridSystemMetadata.SigridSystem import SigridSystem
from SigridSystemMetadata.Users.SigridGetUsersCommand import SigridGetUsersCommand
from SigridSystemMetadata.Users.SigridUser import SigridUser
from Utils.ExcelUtils import set_up_worksheet

MENDIX_ONBOARDING_DELAY=10


@dataclass
class MendixOnboardingOptions:
    pat: Optional[str] = None
    username: Optional[str] = None
    dry_run: bool = False
    output_file: Optional[str] = None


class SigridCustomer:
    def __init__(self, customer: str, token: str, base_url=None):
        self.customer = customer
        self.token = token
        self.base_url = base_url
        # noinspection PyTypeChecker
        self.systems: Dict[str, SigridSystem] = {}
        self.users: Dict[str, SigridUser] = {}

    def get_architecture_JSON_string_for_system(self, system: str):
        return SigridGetArchitectureJSONCommand(self.customer, self.token, system, base_url=self.base_url).do_request()

    def write_architecture_JSON(self, system: str, out_file: str):
        res = self.get_architecture_JSON_string_for_system(system)
        with open(out_file, 'w') as file:
            file.write(res)

    def write_all_architecture_JSONs(self, output_directory: str):
        def get_filename(sys_name):
            return f"{output_directory}/{sys_name}.json"

        if len(self.systems) == 0:
            self.get_available_systems()
        for system in self.systems:
            self.write_architecture_JSON(system, get_filename(system))

    def get_JSON_for_system(self, system: str, date: str = None):
        res = json.loads(SigridGetMaintainabilityJSONCommand(self.customer, self.token, system, date,
                                                             base_url=self.base_url).do_request())
        return res

    def get_available_systems(self):
        res = json.loads(SigridGetMaintainabilityCommand(self.customer, self.token, base_url=self.base_url).do_request())
        systems = res['systems']
        if not isinstance(systems, list):
            systems = [systems]
        for system in systems:
            sys_name = system['system']
            self.systems[sys_name] = (SigridSystem(self.customer, sys_name, self.token, base_url=self.base_url))

    def get_available_systems_with_metadata(self):
        res = json.loads(SigridGetMetadataCommand(self.customer, self.token, base_url=self.base_url).do_request())
        for metadata in res:
            sigrid_system = SigridSystem.from_metadata(metadata, self.token, base_url=self.base_url)
            self.systems[sigrid_system.get_name()] = sigrid_system

    def print_all_systems(self):
        for system in self.systems.values():
            metadata = system.get_metadata().get_data()
            print(metadata)

    def write_objective_csv(self, csv_file):
        with open(csv_file, 'w', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, quotechar="|", quoting=csv.QUOTE_MINIMAL,
                                    fieldnames=['system'] + SigridObjectives.field_order, extrasaction='ignore')
            writer.writeheader()
            for system in self.systems.values():
                writer.writerow({'system': system.get_name()} | system.get_objectives().data)

    def write_metadata_excel(self, xlsx_file):
        wb = Workbook()
        mainsheet = wb.active
        mainsheet.title = f'{self.customer}_Metadata'
        SigridMetadata.write_metadata_header(mainsheet)
        for system in self.systems.values():
            system.get_metadata().write_metadata_to_worksheet(mainsheet)
        SigridMetadata.write_data_validators(mainsheet)
        wb.save(xlsx_file)
        wb.close()

    def write_security_excel(self, xlsx_file, systems=None):

        wb = Workbook()
        mainsheet = wb.active
        mainsheet.title = f'{self.customer}_Security'
        SigridSecurityResults.write_header(mainsheet)


        for system in self.systems.values():
            if systems is None or system.name in systems:
                system.get_security_results().write_metadata_to_worksheet(mainsheet)
        wb.save(xlsx_file)
        wb.close()

    def read_metadata_excel(self, xlsx_file):
        header_row, rows = set_up_worksheet(xlsx_file)
        for row in rows:
            data = [x.value for x in row]
            metadata_dict = dict(zip(header_row, data))
            sys_name = metadata_dict['systemName']
            if sys_name is None:
                print('no system name, skipping')
                continue
            if sys_name not in self.systems:
                self.systems[sys_name] = SigridSystem(self.customer, sys_name, self.token, base_url=self.base_url)
            self.systems[sys_name].set_metadata(metadata_dict)

    def onboard_mendix_from_excel(self, xlsx_file, options: 'MendixOnboardingOptions'):
        header_row, rows = set_up_worksheet(xlsx_file)
        outputs: List[dict] = []

        for row in rows:
            params = self._parse_mendix_row(header_row, row, options)
            if params is None:
                continue
            output = self._onboard_single_system(params, options.dry_run)
            if output is not None:
                outputs.append(output)
            # sleep to avoid overloading the onboarding API
            if not options.dry_run:
                time.sleep(MENDIX_ONBOARDING_DELAY)

        if options.output_file is not None:
            self._write_onboarding_results(outputs, options.output_file)

    @staticmethod
    def _parse_mendix_row(header_row, row, options: 'MendixOnboardingOptions') -> Optional[dict]:
        full_data_dict = dict(zip(header_row, [x.value for x in row]))
        app_name = full_data_dict.get("systemName")
        app_id = full_data_dict.get("externalID") or full_data_dict.get("appId")
        uname = options.username or full_data_dict.get("userName")
        mendix_pat = options.pat or full_data_dict.get("mendixToken")
        team_server_branch = full_data_dict.get("teamServerBranch", '')

        if not (app_name and app_id and mendix_pat):
            # Note: never log mendix_pat - it is a secret.
            print(f"Not onboarding system '{app_name}' (id '{app_id}', user '{uname}'): "
                  f"a required field (systemName, externalID/appId or mendixToken) is missing.")
            return None

        return dict(user_name=uname, mendix_token=mendix_pat, app_id=app_id,
                    app_name=app_name, team_server_branch=team_server_branch)

    def _onboard_single_system(self, params: dict, dry_run: bool) -> Optional[dict]:
        cmd = SigridOnboardQSMCommand(customer=self.customer, token=self.token, base_url=self.base_url, **params)
        result = cmd.do_request(dry_run=dry_run)
        if result is None:
            return None

        app_name = params["app_name"]
        if result.status_code >= 300 or result.status_code < 200:
            print(f'Error: while onboarding system {app_name} got status code {result.status_code} '
                  f'with body {result.content.decode()}')
            return None

        res = result.content.decode()
        print(f'Success: while onboarding system {app_name} got status code {result.status_code} '
              f'with body {res}')
        return json.loads(res)

    def _write_onboarding_results(self, outputs: List[dict], output_file: str):
        wb = Workbook()
        mainsheet = wb.active
        mainsheet.title = f'{self.customer}_Onboarded_Systems'
        mainsheet.append(['customerName', 'systemName'])
        for output in outputs:
            mainsheet.append([output['customerName'], output['systemName']])
        wb.save(output_file)
        wb.close()

    #USER MANAGEMENT

    def load_users_from_sigrid(self) -> Dict[str, SigridUser]:
        cmd = SigridGetUsersCommand(self.customer, self.token, base_url=self.base_url)
        user_list = cmd.get_users()
        return {user.id: user for user in user_list if user is not None}

    def update_users_from_sigrid(self):
        self.users = self.load_users_from_sigrid()

    def update_users_from_dict(self, new_users: Dict[str, SigridUser]):
        self.users.update(new_users)

    def export_users_to_excel(self, xlsx_file):
        if self.users is None or len(self.users) == 0:
            self.users = self.load_users_from_sigrid()

        wb = Workbook()
        mainsheet = wb.active
        mainsheet.title = f'{self.customer}_Users'
        mainsheet.append(SigridUser.field_order)
        for user in self.users.values():
            user.write_to_excel(mainsheet)
        wb.save(xlsx_file)
        wb.close()

    def push_users_to_sigrid(self, dry_run: bool = True):
        for user in self.users.values():
            user.update_permissions(self.customer, self.token, dry_run, base_url=self.base_url)

    def update_sigrid_users_from_excel(self, xlsx_file, dry_run: bool = True):
        self.update_users_from_dict(SigridUser.from_excel(xlsx_file))
        self.push_users_to_sigrid(dry_run)
